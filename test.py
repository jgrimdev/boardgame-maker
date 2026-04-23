import sys
import os
import textwrap
import re
import openpyxl
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QGroupBox, QLabel, QLineEdit, QPushButton, QFileDialog,
                             QTextEdit, QMessageBox, QProgressBar, QSplitter, QSizePolicy)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QIcon, QPixmap

# ==========================================
# ⚙️ KONFIGURACE
# ==========================================
DPI = 300
MM_TO_PX = DPI / 25.4

# Původní rozlišení, na kterém funguje tvůj design (NEMĚNIT)
BASE_W = 500
BASE_H = 700

# ==========================================
# 🎨 STYLY A BARVY KARET
# ==========================================
STYLY_KARET = {
    "pozadi": (255, 255, 255),
    "zbozi": {
        "zelena": (34, 139, 34),
        "zluta": (204, 153, 0),
        "hneda": (139, 69, 19),
        "cervena": (178, 34, 34)
    },
    "lode": {
        "hlavicka_normal": (70, 130, 180),
        "hlavicka_past": (178, 34, 34),
        "text_pasazer": (50, 150, 50),
        "text_past": (200, 0, 0)
    },
    "trh": {
        "hlavicka_pasivni": (70, 130, 180),
        "hlavicka_jednoraz": (178, 34, 34),
        "hlavicka_prestiz": (218, 165, 32),
        "text_cena": (200, 100, 0),
        "text_body": (0, 150, 0)
    }
}


# ==========================================
# 🛠️ POMOCNÉ FUNKCE
# ==========================================
def get_safe_filename(prefix, name):
    safe_name = str(name).lower().replace(" ", "_").replace("'", "").replace("(", "").replace(")", "").replace("/", "_")
    safe_name = safe_name.replace("á", "a").replace("č", "c").replace("é", "e").replace("í", "i")
    safe_name = safe_name.replace("ň", "n").replace("ó", "o").replace("ř", "r").replace("š", "s")
    safe_name = safe_name.replace("ť", "t").replace("ú", "u").replace("ů", "u").replace("ý", "y").replace("ž", "z")
    return f"{prefix}_{safe_name}.png"


def zalom_text(text, sirka_znaku=25):
    if not text: return ""
    return textwrap.fill(str(text), width=sirka_znaku)


def extract_int(text, default=0):
    if text is None or text == "": return default
    if isinstance(text, (int, float)): return int(text)
    s = str(text).replace('–', '-').replace('—', '-').strip()
    try:
        return int(s)
    except ValueError:
        pass
    hodnota = ''.join(c for c in s if c.isdigit() or c == '-')
    if not hodnota or hodnota == '-': return default
    try:
        return int(hodnota)
    except ValueError:
        return default


# ==========================================
# 🖌️ VYKRESLOVACÍ JÁDRO (Zajišťuje 100% shodu původního vzhledu)
# ==========================================
class CardRenderer:
    def __init__(self):
        try:
            self.font_big = ImageFont.truetype("arialbd.ttf", 150)
            self.font_mid = ImageFont.truetype("arialbd.ttf", 50)
            self.font_small = ImageFont.truetype("arial.ttf", 25)
            self.font_title = ImageFont.truetype("arialbd.ttf", 35)
        except IOError:
            self.font_big = self.font_mid = self.font_small = self.font_title = ImageFont.load_default()

    def nakresli_ramecek(self, draw, tloustka=3, barva=(0, 0, 0)):
        draw.rectangle([0, 0, BASE_W - 1, BASE_H - 1], outline=barva, width=tloustka)

    def vycentruj_text(self, draw, text, font, y_pos, barva=(0, 0, 0)):
        try:
            bbox = draw.textbbox((0, 0), str(text), font=font)
            sirka_textu = bbox[2] - bbox[0]
        except AttributeError:
            sirka_textu, _ = draw.textsize(str(text), font=font)
        x_pos = (BASE_W - sirka_textu) / 2
        draw.text((x_pos, y_pos), str(text), font=font, fill=barva)

    def vycentruj_text_v_boxu(self, draw, text, font, box_x1, box_x2, y_pos, barva=(0, 0, 0)):
        try:
            bbox = draw.textbbox((0, 0), str(text), font=font)
            sirka_textu = bbox[2] - bbox[0]
        except AttributeError:
            sirka_textu, _ = draw.textsize(str(text), font=font)
        stred_boxu = box_x1 + (box_x2 - box_x1) / 2
        x_pos = stred_boxu - (sirka_textu / 2)
        draw.text((x_pos, y_pos), str(text), font=font, fill=barva)

    def render(self, card_data, target_w_px, target_h_px):
        """Vykreslí kartu na 500x700 a pak ji zvětší/zmenší na požadovaný rozměr"""
        img = Image.new('RGB', (BASE_W, BASE_H), STYLY_KARET["pozadi"])
        draw = ImageDraw.Draw(img)

        kat = card_data.get("kategorie", "")

        if kat in ["ZBOŽÍ", "ZBOZI"]:
            self._kresli_zbozi(draw, card_data)
        elif kat in ["LOĎ", "LOD"]:
            self._kresli_lod(draw, card_data)
        elif kat == "TRH":
            self._kresli_trh(draw, card_data)
        else:
            self._kresli_ostatni(draw, card_data)

        # Fáze škálování do tiskového rozlišení
        if target_w_px != BASE_W or target_h_px != BASE_H:
            img = img.resize((int(target_w_px), int(target_h_px)), Image.Resampling.LANCZOS)

        return img

    def _kresli_zbozi(self, draw, c):
        cislo = c["cislo"]
        nazev = c["nazev"]

        if cislo <= 15:
            barva = STYLY_KARET["zbozi"]["zelena"]
        elif cislo <= 30:
            barva = STYLY_KARET["zbozi"]["zluta"]
        elif cislo <= 45:
            barva = STYLY_KARET["zbozi"]["hneda"]
        else:
            barva = STYLY_KARET["zbozi"]["cervena"]

        self.nakresli_ramecek(draw, tloustka=5, barva=barva)
        draw.rectangle([15, 15, BASE_W - 15, BASE_H - 15], outline=barva, width=2)
        draw.text((40, 40), str(cislo), font=self.font_mid, fill=barva)
        draw.text((BASE_W - 100, BASE_H - 90), str(cislo), font=self.font_mid, fill=barva)
        self.vycentruj_text(draw, str(cislo), self.font_big, BASE_H / 2 - 90, barva=barva)
        self.vycentruj_text(draw, str(nazev).upper(), self.font_title, BASE_H - 150, barva=(0, 0, 0))
        self.vycentruj_text(draw, "ZBOŽÍ", self.font_small, BASE_H - 100, barva=(100, 100, 100))

    def _kresli_lod(self, draw, c):
        self.nakresli_ramecek(draw)
        je_past = c["je_past"]
        kapacita = max(1, c["kapacita"])

        barva_hlavicky = STYLY_KARET["lode"]["hlavicka_past"] if je_past else STYLY_KARET["lode"]["hlavicka_normal"]
        draw.rectangle([10, 10, BASE_W - 10, 110], outline=barva_hlavicky, width=4)
        self.vycentruj_text(draw, c["nazev"], self.font_title, 25, barva=barva_hlavicky)
        podtitul = "PAST" if je_past else "LOĎ"
        self.vycentruj_text(draw, podtitul, self.font_small, 70, barva=barva_hlavicky)

        start_y = 150;
        mezera = 65
        self.vycentruj_text(draw, f"KAPACITA: {kapacita}", self.font_small, start_y - 35)

        for i in range(kapacita):
            y = start_y + (i * mezera)
            draw.rectangle([80, y, BASE_W - 80, y + 50], outline=(0, 0, 0), width=2)
            ikona = "[ CÍL ]" if i == kapacita - 1 else "[ zboží ]"
            self.vycentruj_text(draw, ikona, self.font_small, y + 10, barva=(100, 100, 100))

        y_odmeny = start_y + (kapacita * mezera) + 30
        draw.rectangle([30, y_odmeny, BASE_W - 30, BASE_H - 30], outline=(0, 0, 0), width=3)
        self.vycentruj_text(draw, "Poslední zboží:", self.font_small, y_odmeny + 15)

        barva_pasazer = STYLY_KARET["lode"]["text_past"] if je_past else STYLY_KARET["lode"]["text_pasazer"]
        self.vycentruj_text(draw, f"Poslední karta: {c['body']} VB", self.font_title, y_odmeny + 60, barva=(0, 0, 0))
        self.vycentruj_text(draw, f"Náklad zboží: {c['mince']} Mušle", self.font_mid, y_odmeny + 110,
                            barva=barva_pasazer)

        efekt = c.get("efekt", "")
        if efekt and str(efekt).strip() != "-":
            zalomene = zalom_text(efekt, 35)
            draw.multiline_text((50, y_odmeny + 180), zalomene, font=self.font_small, fill=(200, 0, 0), align="center",
                                spacing=5)

    def _kresli_trh(self, draw, c):
        self.nakresli_ramecek(draw)
        podtyp = str(c["podtyp"]).lower()

        if "pasiv" in podtyp:
            barva_hlavicky = STYLY_KARET["trh"]["hlavicka_pasivni"]
        elif "jednor" in podtyp or "trik" in podtyp:
            barva_hlavicky = STYLY_KARET["trh"]["hlavicka_jednoraz"]
        else:
            barva_hlavicky = STYLY_KARET["trh"]["hlavicka_prestiz"]

        draw.rectangle([10, 10, BASE_W - 10, 110], outline=barva_hlavicky, width=4)
        self.vycentruj_text(draw, c["nazev"], self.font_title, 25, barva=barva_hlavicky)
        self.vycentruj_text(draw, str(c["podtyp"]).upper(), self.font_small, 70, barva=barva_hlavicky)

        y_stats = 150
        draw.rectangle([50, y_stats, 220, y_stats + 100], outline=(0, 0, 0), width=3)
        self.vycentruj_text_v_boxu(draw, "CENA", self.font_small, 50, 220, y_stats + 10)
        cena_txt = str(c["cena"]) if (c["cena"] and str(c["cena"]) not in ["0", "-"]) else "-"
        self.vycentruj_text_v_boxu(draw, cena_txt, self.font_mid, 50, 220, y_stats + 40,
                                   barva=STYLY_KARET["trh"]["text_cena"])

        draw.rectangle([BASE_W - 220, y_stats, BASE_W - 50, y_stats + 100], outline=(0, 0, 0), width=3)
        self.vycentruj_text_v_boxu(draw, "BODY (VB)", self.font_small, BASE_W - 220, BASE_W - 50, y_stats + 10)
        body_txt = f"+{c['body']}" if (c["body"] and str(c["body"]) not in ["0", "-"]) else "-"
        self.vycentruj_text_v_boxu(draw, body_txt, self.font_mid, BASE_W - 220, BASE_W - 50, y_stats + 40,
                                   barva=STYLY_KARET["trh"]["text_body"])

        efekt = c.get("efekt", "")
        if efekt and str(efekt).strip() != "-":
            draw.rectangle([30, y_stats + 140, BASE_W - 30, BASE_H - 40], outline=(100, 100, 100), width=2)
            self.vycentruj_text(draw, "EFEKT KARTY:", self.font_small, y_stats + 160)
            zalomene = zalom_text(efekt, 22)
            draw.multiline_text((50, y_stats + 220), zalomene, font=self.font_title, fill=(0, 0, 0), align="center",
                                spacing=10)

    def _kresli_ostatni(self, draw, c):
        self.nakresli_ramecek(draw)
        barva_hlavicky = (100, 100, 100)
        draw.rectangle([10, 10, BASE_W - 10, 110], outline=barva_hlavicky, width=4)
        self.vycentruj_text(draw, str(c["nazev"]).upper(), self.font_title, 35, barva=barva_hlavicky)

        efekt = c.get("efekt", "")
        if efekt and str(efekt).strip() != "-":
            zalomene = zalom_text(efekt, 25)
            draw.multiline_text((40, 150), zalomene, font=self.font_title, fill=(0, 0, 0), align="left", spacing=15)


# ==========================================
# ⚙️ VLÁKNO GENERÁTORU (Ukládání)
# ==========================================
class CardGeneratorWorker(QThread):
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, card_data_list, dir_out, excel_out, width_px, height_px):
        super().__init__()
        self.card_data_list = card_data_list
        self.dir_out = dir_out
        self.excel_out = excel_out
        self.w_px = width_px
        self.h_px = height_px
        self.renderer = CardRenderer()

    def run(self):
        try:
            os.makedirs(self.dir_out, exist_ok=True)
            self.log_signal.emit("\n--- ZAČÍNÁM GENEROVAT OBRÁZKY ---")

            tisk_data = []
            total_cards = len(self.card_data_list)

            for i, card in enumerate(self.card_data_list):
                self.progress_signal.emit(int(((i + 1) / total_cards) * 100))

                # Generování PIL obrázku
                img = self.renderer.render(card, self.w_px, self.h_px)

                # Sestavení jména a logu
                kat = card.get("kategorie", "Ostatni")
                if kat in ["ZBOŽÍ", "ZBOZI"]:
                    filename = get_safe_filename("zbozi", f'{card["nazev"]}_{card["cislo"]}')
                elif kat in ["LOĎ", "LOD"]:
                    filename = get_safe_filename("lod", card["nazev"])
                elif kat == "TRH":
                    filename = get_safe_filename("trh", card["nazev"])
                else:
                    filename = get_safe_filename("ostatni", card["nazev"])

                # Uložení na disk
                img.save(os.path.join(self.dir_out, filename))
                tisk_data.append([card["zobrazovany_nazev"], kat, filename, card["pocet"]])

                self.log_signal.emit(f"Uloženo: {filename}")

            # Uložení Excelu pro sazbu
            self.log_signal.emit("\n💾 Vytvářím tiskový Excel...")
            out_wb = openpyxl.Workbook()
            out_sheet = out_wb.active
            out_sheet.title = "Data_pro_sazbu"
            out_sheet.append(["Název Karty", "Typ Karty", "Soubor_Obrazku", "Mnozstvi"])
            for radek in tisk_data:
                out_sheet.append(radek)
            out_wb.save(self.excel_out)

            self.finished_signal.emit(True, "Generování všech karet proběhlo úspěšně!")

        except Exception as e:
            self.finished_signal.emit(False, str(e))


# ==========================================
# 🖥️ GRAFICKÉ ROZHRANÍ (PyQt6)
# ==========================================
class CardGeneratorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Generátor Karet (Data -> Obrázky)")
        self.resize(1100, 750)
        self.worker = None

        # Uchovává v paměti rozparsovaná data pro bleskový náhled
        self.parsed_cards = []
        self.preview_idx = 0
        self.renderer = CardRenderer()

        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.setCentralWidget(self.splitter)

        self.setup_ui()

    def setup_ui(self):
        # --- LEVÝ PANEL (Ovládání) ---
        left_widget = QWidget()
        main_layout = QVBoxLayout(left_widget)

        # 1. ZDROJE
        gb_sources = QGroupBox("1. Zdroje dat")
        src_layout = QVBoxLayout()

        row_in = QHBoxLayout()
        self.inp_excel_in = QLineEdit("karty_data.xlsx")
        btn_in = QPushButton("Vybrat Data")
        btn_in.clicked.connect(self.select_in_file)
        row_in.addWidget(QLabel("Vstupní Excel:"))
        row_in.addWidget(self.inp_excel_in)
        row_in.addWidget(btn_in)
        src_layout.addLayout(row_in)

        # Načítací tlačítko, které naplní paměť daty z Excelu
        self.btn_load_data = QPushButton("Načíst data z Excelu (Aktualizovat)")
        self.btn_load_data.setStyleSheet("background-color: #ffd700; font-weight: bold;")
        self.btn_load_data.clicked.connect(self.parse_excel_data)
        src_layout.addWidget(self.btn_load_data)

        gb_sources.setLayout(src_layout)
        main_layout.addWidget(gb_sources)

        # 2. VÝSTUP A NASTAVENÍ
        gb_settings = QGroupBox("2. Nastavení exportu")
        settings_layout = QVBoxLayout()

        row_out_dir = QHBoxLayout()
        self.inp_dir_out = QLineEdit("karty_export")
        btn_dir = QPushButton("Složka")
        btn_dir.clicked.connect(self.select_out_dir)
        row_out_dir.addWidget(QLabel("Obrázky do:"))
        row_out_dir.addWidget(self.inp_dir_out)
        row_out_dir.addWidget(btn_dir)
        settings_layout.addLayout(row_out_dir)

        row_out_excel = QHBoxLayout()
        self.inp_excel_out = QLineEdit("tiskova_data_pro_sazbu.xlsx")
        row_out_excel.addWidget(QLabel("Tiskový Excel:"))
        row_out_excel.addWidget(self.inp_excel_out)
        settings_layout.addLayout(row_out_excel)

        row_dims = QHBoxLayout()
        self.inp_width = QLineEdit("63")
        self.inp_height = QLineEdit("88")
        row_dims.addWidget(QLabel("Šířka (mm):"))
        row_dims.addWidget(self.inp_width)
        row_dims.addWidget(QLabel("Výška (mm):"))
        row_dims.addWidget(self.inp_height)
        settings_layout.addLayout(row_dims)

        gb_settings.setLayout(settings_layout)
        main_layout.addWidget(gb_settings)

        # 3. AKCE
        self.btn_generate = QPushButton("Generovat a uložit všechny karty")
        self.btn_generate.setStyleSheet("background-color: #2e7d32; color: white; font-weight: bold; padding: 10px;")
        self.btn_generate.setEnabled(False)  # Povolí se po načtení dat
        self.btn_generate.clicked.connect(self.start_generation)
        main_layout.addWidget(self.btn_generate)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        main_layout.addWidget(self.progress)

        self.console = QTextEdit()
        self.console.setReadOnly(True)
        self.console.setStyleSheet("background-color: #1e1e1e; color: #00ff00; font-family: Consolas;")
        main_layout.addWidget(self.console)

        self.splitter.addWidget(left_widget)

        # --- PRAVÝ PANEL (Náhled s Navigací) ---
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_widget.setStyleSheet("background-color: #404040;")

        # Ovládání náhledu
        nav_layout = QHBoxLayout()
        nav_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.btn_prev = QPushButton("◄ Předchozí")
        self.btn_prev.setFixedWidth(100)
        self.btn_prev.clicked.connect(self.preview_prev)

        self.lbl_nav = QLabel("0 / 0")
        self.lbl_nav.setStyleSheet("color: white; font-weight: bold; font-size: 14px;")
        self.lbl_nav.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_nav.setFixedWidth(100)

        self.btn_next = QPushButton("Další ►")
        self.btn_next.setFixedWidth(100)
        self.btn_next.clicked.connect(self.preview_next)

        nav_layout.addWidget(self.btn_prev)
        nav_layout.addWidget(self.lbl_nav)
        nav_layout.addWidget(self.btn_next)
        right_layout.addLayout(nav_layout)

        # Samotný obrázek náhledu
        self.lbl_preview = QLabel("Načtěte data z Excelu pro zobrazení náhledu.")
        self.lbl_preview.setStyleSheet("color: white;")
        self.lbl_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_preview.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        right_layout.addWidget(self.lbl_preview)

        self.splitter.addWidget(right_widget)
        self.splitter.setSizes([450, 650])

    def select_in_file(self):
        f, _ = QFileDialog.getOpenFileName(self, "Vyberte vstupní Excel", "", "Excel (*.xlsx *.xls)")
        if f: self.inp_excel_in.setText(f)

    def select_out_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Vyberte složku pro obrázky")
        if d: self.inp_dir_out.setText(d)

    def log_msg(self, msg):
        self.console.append(msg)
        scrollbar = self.console.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    # ==========================================
    # 🧠 PARSOVÁNÍ EXCELU (Proběhne jen jednou)
    # ==========================================
    def parse_excel_data(self):
        excel_in = self.inp_excel_in.text()
        if not os.path.exists(excel_in):
            QMessageBox.critical(self, "Chyba", f"Soubor '{excel_in}' nebyl nalezen!")
            return

        self.console.clear()
        self.log_msg(f"📚 Čtu strukturu Excelu '{excel_in}'...")
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        self.parsed_cards = []

        try:
            wb = openpyxl.load_workbook(excel_in, data_only=True)
            sheet = wb.active

            col_map = {"kat": 0, "pod": 1, "nazev": 2, "kap": 4, "cena": 5, "body": 6, "efekt": 7, "pocet": 8}
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                for i, h in enumerate(row):
                    if not h: continue
                    h_lower = str(h).lower()
                    if "kateg" in h_lower:
                        col_map["kat"] = i
                    elif "pod-typ" in h_lower or "podtyp" in h_lower:
                        col_map["pod"] = i
                    elif "oceán" in h_lower or "ocean" in h_lower:
                        col_map["nazev"] = i
                    elif "kapacita" in h_lower or "čísla" in h_lower or "cisla" in h_lower:
                        col_map["kap"] = i
                    elif "cena" in h_lower or "odměna" in h_lower or "odmena" in h_lower:
                        col_map["cena"] = i
                    elif "body" in h_lower or "vb" in h_lower:
                        col_map["body"] = i
                    elif "efekt" in h_lower or "text" in h_lower:
                        col_map["efekt"] = i
                    elif "počet" in h_lower or "pocet" in h_lower or "ks" in h_lower:
                        col_map["pocet"] = i
                break

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or row[col_map["kat"]] is None: continue

                def get_val(key):
                    idx = col_map[key]
                    return str(row[idx]).strip() if len(row) > idx and row[idx] is not None else ""

                kategorie = get_val("kat").upper()
                nazev = get_val("nazev")
                pocet = extract_int(get_val("pocet"), 1)

                base_card = {
                    "kategorie": kategorie,
                    "podtyp": get_val("pod"),
                    "nazev": nazev,
                    "zobrazovany_nazev": nazev,
                    "kapacita": extract_int(get_val("kap"), 0),
                    "kapacita_str": get_val("kap"),
                    "cena": extract_int(get_val("cena"), 0),
                    "mince": extract_int(get_val("cena"), 0),
                    "body": extract_int(get_val("body"), 0),
                    "efekt": get_val("efekt"),
                    "je_past": ("past" in get_val("pod").lower()),
                    "pocet": pocet
                }

                if kategorie in ["ZBOŽÍ", "ZBOZI"]:
                    cisla_rozsah = re.findall(r'\d+', base_card["kapacita_str"])
                    if len(cisla_rozsah) >= 2:
                        for i in range(int(cisla_rozsah[0]), int(cisla_rozsah[1]) + 1):
                            new_card = base_card.copy()
                            new_card["cislo"] = i
                            new_card["zobrazovany_nazev"] = f"{nazev} (Číslo {i})"
                            new_card["pocet"] = 1
                            self.parsed_cards.append(new_card)
                else:
                    self.parsed_cards.append(base_card)

            self.log_msg(f"✅ Načteno celkem {len(self.parsed_cards)} unikátních karet do paměti.")

            if self.parsed_cards:
                self.btn_generate.setEnabled(True)
                self.preview_idx = 0
                self.update_preview()
            else:
                self.lbl_nav.setText("0 / 0")

        except Exception as e:
            QMessageBox.critical(self, "Chyba čtení", str(e))
        finally:
            QApplication.restoreOverrideCursor()

    # ==========================================
    # 🖼️ BLESKOVÝ NÁHLED (Okamžité zobrazení)
    # ==========================================
    def get_dimensions_px(self):
        try:
            w_mm = float(self.inp_width.text().replace(',', '.'))
            h_mm = float(self.inp_height.text().replace(',', '.'))
            return int(w_mm * MM_TO_PX), int(h_mm * MM_TO_PX)
        except ValueError:
            return None, None

    def preview_prev(self):
        if self.parsed_cards and self.preview_idx > 0:
            self.preview_idx -= 1
            self.update_preview()

    def preview_next(self):
        if self.parsed_cards and self.preview_idx < len(self.parsed_cards) - 1:
            self.preview_idx += 1
            self.update_preview()

    def update_preview(self):
        if not self.parsed_cards: return
        w_px, h_px = self.get_dimensions_px()
        if w_px is None: return

        self.lbl_nav.setText(f"{self.preview_idx + 1} / {len(self.parsed_cards)}")

        # Generování pouze jedné karty rovnou v hlavním vlákně (je to super rychlé)
        card = self.parsed_cards[self.preview_idx]
        pil_img = self.renderer.render(card, w_px, h_px)

        # Zobrazení do QLabelu
        bytes_io = BytesIO()
        pil_img.save(bytes_io, format="PNG")
        pixmap = QPixmap()
        pixmap.loadFromData(bytes_io.getvalue())

        pw = self.lbl_preview.width()
        ph = self.lbl_preview.height()
        if pw > 20 and ph > 20:
            scaled_pixmap = pixmap.scaled(pw - 20, ph - 20, Qt.AspectRatioMode.KeepAspectRatio,
                                          Qt.TransformationMode.SmoothTransformation)
            self.lbl_preview.setPixmap(scaled_pixmap)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_preview()

    # ==========================================
    # 🚀 EXPORT VŠECH KARET
    # ==========================================
    def start_generation(self):
        if not self.parsed_cards:
            QMessageBox.warning(self, "Upozornění", "Nejprve načtěte data z Excelu.")
            return

        w_px, h_px = self.get_dimensions_px()
        if w_px is None:
            QMessageBox.critical(self, "Chyba", "Zadejte správně milimetry.")
            return

        self.btn_generate.setEnabled(False)
        self.btn_load_data.setEnabled(False)
        self.progress.setValue(0)

        # Spuštění vlákna
        self.worker = CardGeneratorWorker(
            card_data_list=self.parsed_cards,
            dir_out=self.inp_dir_out.text(),
            excel_out=self.inp_excel_out.text(),
            width_px=w_px, height_px=h_px
        )
        self.worker.log_signal.connect(self.log_msg)
        self.worker.progress_signal.connect(self.progress.setValue)
        self.worker.finished_signal.connect(self.generation_finished)
        self.worker.start()

    def generation_finished(self, success, msg):
        self.progress.setValue(100)
        self.btn_generate.setEnabled(True)
        self.btn_load_data.setEnabled(True)
        if success:
            self.log_msg(f"\n✅ {msg}")
            QMessageBox.information(self, "Hotovo", msg)
        else:
            self.log_msg(f"\n❌ CHYBA: {msg}")
            QMessageBox.critical(self, "Chyba", msg)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = CardGeneratorApp()
    window.show()
    sys.exit(app.exec())