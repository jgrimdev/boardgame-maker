import sys
import os
import re
import openpyxl
from io import BytesIO

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QGroupBox, QLabel, QLineEdit, QPushButton, QFileDialog,
                             QTextEdit, QMessageBox, QProgressBar, QSplitter, QSizePolicy, QSlider, QComboBox)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QIcon, QPixmap

# IMPORT VLASTNÍHO RENDERERU (Díky tomuhle se aplikace úplně oddělí od grafiky karet)
from card_renderer import CardRenderer, get_safe_filename

DPI = 300
MM_TO_PX = DPI / 25.4


def extract_int(text, default=0):
    if text is None or text == "": return default
    if isinstance(text, (int, float)): return int(text)
    s = str(text).replace('–', '-').replace('—', '-').strip()
    try:
        return int(s)
    except ValueError:
        pass
    hodnota = ''.join(c for c in s if c.isdigit() or c == '-')
    if not hodnota or hodnota == '-': return default
    try:
        return int(hodnota)
    except ValueError:
        return default


# ==========================================
# ⚙️ VLÁKNO GENERÁTORU (Ukládání na disk)
# ==========================================
class CardGeneratorWorker(QThread):
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, card_data_list, dir_out, excel_out, width_px, height_px):
        super().__init__()
        self.card_data_list = card_data_list
        self.dir_out = dir_out
        self.excel_out = excel_out
        self.w_px = width_px
        self.h_px = height_px
        self.renderer = CardRenderer()

    def run(self):
        try:
            os.makedirs(self.dir_out, exist_ok=True)
            self.log_signal.emit("\n--- ZAČÍNÁM MASOVÝ EXPORT OBRÁZKŮ ---")

            tisk_data = []
            total_cards = len(self.card_data_list)

            for i, card in enumerate(self.card_data_list):
                self.progress_signal.emit(int(((i + 1) / total_cards) * 100))

                img = self.renderer.render(card, self.w_px, self.h_px)
                filename = card["filename"]

                img.save(os.path.join(self.dir_out, filename))
                tisk_data.append([card["zobrazovany_nazev"], card["kategorie"], filename, card["pocet"]])

                # Názvy souborů zpět v konzoli!
                self.log_signal.emit(f"Uloženo: {filename}")

            self.log_signal.emit("\n💾 Vytvářím tiskový Excel pro další zpracování...")
            out_wb = openpyxl.Workbook()
            out_sheet = out_wb.active
            out_sheet.title = "Data_pro_sazbu"
            out_sheet.append(["Název Karty", "Typ Karty", "Soubor_Obrazku", "Mnozstvi"])
            for radek in tisk_data:
                out_sheet.append(radek)
            out_wb.save(self.excel_out)

            self.finished_signal.emit(True, "Generování všech karet proběhlo úspěšně!")

        except Exception as e:
            self.finished_signal.emit(False, str(e))


# ==========================================
# 🖥️ GRAFICKÉ ROZHRANÍ (PyQt6)
# ==========================================
class CardGeneratorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Generátor Karet (Data -> Obrázky)")
        self.resize(1150, 750)
        self.worker = None

        self.parsed_cards = []
        self.filtered_cards = []  # Půjde listovat jen přes vybranou kategorii
        self.preview_idx = 0
        self.renderer = CardRenderer()

        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.setCentralWidget(self.splitter)

        self.setup_ui()

    def setup_ui(self):
        # --- LEVÝ PANEL ---
        left_widget = QWidget()
        main_layout = QVBoxLayout(left_widget)

        gb_sources = QGroupBox("1. Zdroje dat")
        src_layout = QVBoxLayout()
        row_in = QHBoxLayout()
        self.inp_excel_in = QLineEdit("karty_data.xlsx")
        btn_in = QPushButton("Vybrat Data")
        btn_in.clicked.connect(self.select_in_file)
        row_in.addWidget(QLabel("Vstupní Excel:"))
        row_in.addWidget(self.inp_excel_in)
        row_in.addWidget(btn_in)
        src_layout.addLayout(row_in)

        self.btn_load_data = QPushButton("Načíst data z Excelu (Aktualizovat)")
        self.btn_load_data.setStyleSheet("background-color: #ffd700; font-weight: bold;")
        self.btn_load_data.clicked.connect(self.parse_excel_data)
        src_layout.addWidget(self.btn_load_data)

        gb_sources.setLayout(src_layout)
        main_layout.addWidget(gb_sources)

        gb_settings = QGroupBox("2. Nastavení exportu")
        settings_layout = QVBoxLayout()
        row_out_dir = QHBoxLayout()
        self.inp_dir_out = QLineEdit("karty_export")
        btn_dir = QPushButton("Složka")
        btn_dir.clicked.connect(self.select_out_dir)
        row_out_dir.addWidget(QLabel("Obrázky do:"))
        row_out_dir.addWidget(self.inp_dir_out)
        row_out_dir.addWidget(btn_dir)
        settings_layout.addLayout(row_out_dir)

        row_out_excel = QHBoxLayout()
        self.inp_excel_out = QLineEdit("tiskova_data_pro_sazbu.xlsx")
        row_out_excel.addWidget(QLabel("Tiskový Excel:"))
        row_out_excel.addWidget(self.inp_excel_out)
        settings_layout.addLayout(row_out_excel)

        row_dims = QHBoxLayout()
        self.inp_width = QLineEdit("63")
        self.inp_height = QLineEdit("88")
        row_dims.addWidget(QLabel("Cílová šířka (mm):"))
        row_dims.addWidget(self.inp_width)
        row_dims.addWidget(QLabel("Výška (mm):"))
        row_dims.addWidget(self.inp_height)
        settings_layout.addLayout(row_dims)

        gb_settings.setLayout(settings_layout)
        main_layout.addWidget(gb_settings)

        self.btn_generate = QPushButton("Generovat a uložit všechny karty")
        self.btn_generate.setStyleSheet("background-color: #2e7d32; color: white; font-weight: bold; padding: 10px;")
        self.btn_generate.setEnabled(False)
        self.btn_generate.clicked.connect(self.start_generation)
        main_layout.addWidget(self.btn_generate)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        main_layout.addWidget(self.progress)

        self.console = QTextEdit()
        self.console.setReadOnly(True)
        self.console.setStyleSheet("background-color: #1e1e1e; color: #00ff00; font-family: Consolas;")
        main_layout.addWidget(self.console)

        self.splitter.addWidget(left_widget)

        # --- PRAVÝ PANEL (Náhled s filtry) ---
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_widget.setStyleSheet("background-color: #404040;")

        # Filtr kategorií
        filter_layout = QHBoxLayout()
        filter_lbl = QLabel("Zobrazit kategorii:")
        filter_lbl.setStyleSheet("color: white;")
        self.combo_filter = QComboBox()
        self.combo_filter.setEnabled(False)
        self.combo_filter.currentTextChanged.connect(self.filter_changed)
        filter_layout.addWidget(filter_lbl)
        filter_layout.addWidget(self.combo_filter)
        filter_layout.addStretch()
        right_layout.addLayout(filter_layout)

        self.slider_nav = QSlider(Qt.Orientation.Horizontal)
        self.slider_nav.setMinimum(0)
        self.slider_nav.setEnabled(False)
        self.slider_nav.valueChanged.connect(self.slider_changed)
        right_layout.addWidget(self.slider_nav)

        nav_layout = QHBoxLayout()
        nav_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.btn_prev = QPushButton("◄")
        self.btn_prev.setFixedWidth(50)
        self.btn_prev.clicked.connect(self.preview_prev)

        self.inp_nav = QLineEdit("0")
        self.inp_nav.setFixedWidth(50)
        self.inp_nav.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.inp_nav.setEnabled(False)
        self.inp_nav.returnPressed.connect(self.input_changed)

        self.lbl_nav_total = QLabel("/ 0")
        self.lbl_nav_total.setStyleSheet("color: white; font-weight: bold; font-size: 14px;")

        self.btn_next = QPushButton("►")
        self.btn_next.setFixedWidth(50)
        self.btn_next.clicked.connect(self.preview_next)

        nav_layout.addWidget(self.btn_prev)
        nav_layout.addWidget(self.inp_nav)
        nav_layout.addWidget(self.lbl_nav_total)
        nav_layout.addWidget(self.btn_next)
        right_layout.addLayout(nav_layout)

        self.lbl_preview = QLabel("Načtěte data z Excelu pro zobrazení náhledu.")
        self.lbl_preview.setStyleSheet("color: white;")
        self.lbl_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_preview.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        right_layout.addWidget(self.lbl_preview)

        self.splitter.addWidget(right_widget)
        self.splitter.setSizes([450, 700])

    def select_in_file(self):
        f, _ = QFileDialog.getOpenFileName(self, "Vyberte vstupní Excel", "", "Excel (*.xlsx *.xls)")
        if f: self.inp_excel_in.setText(f)

    def select_out_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Vyberte složku pro obrázky")
        if d: self.inp_dir_out.setText(d)

    def log_msg(self, msg):
        self.console.append(msg)
        scrollbar = self.console.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    # ==========================================
    # 🧠 PARSOVÁNÍ EXCELU
    # ==========================================
    def parse_excel_data(self):
        excel_in = self.inp_excel_in.text()
        if not os.path.exists(excel_in):
            QMessageBox.critical(self, "Chyba", f"Soubor '{excel_in}' nebyl nalezen!")
            return

        self.console.clear()
        self.log_msg(f"📚 Čtu strukturu Excelu '{excel_in}'...\n")
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        self.parsed_cards = []

        try:
            wb = openpyxl.load_workbook(excel_in, data_only=True)
            sheet = wb.active

            col_map = {"kat": 0, "pod": 1, "nazev": 2, "kap": 4, "cena": 5, "body": 6, "efekt": 7, "pocet": 8}
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                for i, h in enumerate(row):
                    if not h: continue
                    h_lower = str(h).lower()
                    if "kateg" in h_lower:
                        col_map["kat"] = i
                    elif "pod-typ" in h_lower or "podtyp" in h_lower:
                        col_map["pod"] = i
                    elif "oceán" in h_lower or "ocean" in h_lower:
                        col_map["nazev"] = i
                    elif "kapacita" in h_lower or "čísla" in h_lower or "cisla" in h_lower:
                        col_map["kap"] = i
                    elif "cena" in h_lower or "odměna" in h_lower or "odmena" in h_lower:
                        col_map["cena"] = i
                    elif "body" in h_lower or "vb" in h_lower:
                        col_map["body"] = i
                    elif "efekt" in h_lower or "text" in h_lower:
                        col_map["efekt"] = i
                    elif "počet" in h_lower or "pocet" in h_lower or "ks" in h_lower:
                        col_map["pocet"] = i
                break

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or row[col_map["kat"]] is None: continue

                def get_val(key):
                    idx = col_map[key]
                    return str(row[idx]).strip() if len(row) > idx and row[idx] is not None else ""

                kategorie = get_val("kat").upper()
                nazev = get_val("nazev")
                pocet = extract_int(get_val("pocet"), 1)

                base_card = {
                    "kategorie": kategorie,
                    "podtyp": get_val("pod"),
                    "nazev": nazev,
                    "zobrazovany_nazev": nazev,
                    "kapacita": extract_int(get_val("kap"), 0),
                    "kapacita_str": get_val("kap"),
                    "cena": extract_int(get_val("cena"), 0),
                    "mince": extract_int(get_val("cena"), 0),
                    "body": extract_int(get_val("body"), 0),
                    "efekt": get_val("efekt"),
                    "je_past": ("past" in get_val("pod").lower()),
                    "pocet": pocet,
                    "filename": ""  # Přidáno rovnou při parsování
                }

                if kategorie in ["ZBOŽÍ", "ZBOZI"]:
                    cisla_rozsah = re.findall(r'\d+', base_card["kapacita_str"])
                    if len(cisla_rozsah) >= 2:
                        for i in range(int(cisla_rozsah[0]), int(cisla_rozsah[1]) + 1):
                            new_card = base_card.copy()
                            new_card["cislo"] = i
                            new_card["zobrazovany_nazev"] = f"{nazev} (Číslo {i})"
                            new_card["pocet"] = 1
                            new_card["filename"] = get_safe_filename("zbozi", f"{nazev}_{i}")
                            self.parsed_cards.append(new_card)
                            self.log_msg(f"Načteno -> {new_card['filename']}")
                else:
                    if kategorie in ["LOĎ", "LOD"]:
                        base_card["filename"] = get_safe_filename("lod", nazev)
                    elif kategorie == "TRH":
                        base_card["filename"] = get_safe_filename("trh", nazev)
                    else:
                        base_card["filename"] = get_safe_filename("ostatni", nazev)
                    self.parsed_cards.append(base_card)
                    self.log_msg(f"Načteno -> {base_card['filename']}")

            self.log_msg(f"\n✅ ÚSPĚCH: Načteno celkem {len(self.parsed_cards)} unikátních karet do paměti.")

            if self.parsed_cards:
                self.btn_generate.setEnabled(True)

                # Zjištění unikátních kategorií pro filtr
                kategorie_set = set(c["kategorie"] for c in self.parsed_cards)
                self.combo_filter.blockSignals(True)
                self.combo_filter.clear()
                self.combo_filter.addItem("Všechny karty")
                self.combo_filter.addItems(sorted(kategorie_set))
                self.combo_filter.blockSignals(False)
                self.combo_filter.setEnabled(True)

                # Založení na "Všechny karty"
                self.filter_changed("Všechny karty")
            else:
                self.lbl_nav_total.setText("/ 0")

        except Exception as e:
            QMessageBox.critical(self, "Chyba čtení", str(e))
        finally:
            QApplication.restoreOverrideCursor()

    # ==========================================
    # 🖼️ NÁHLED A FILTROVÁNÍ
    # ==========================================
    def filter_changed(self, kat_name):
        if not self.parsed_cards: return

        if kat_name == "Všechny karty":
            self.filtered_cards = self.parsed_cards
        else:
            self.filtered_cards = [c for c in self.parsed_cards if c["kategorie"] == kat_name]

        if self.filtered_cards:
            self.slider_nav.setEnabled(True)
            self.inp_nav.setEnabled(True)
            self.slider_nav.setMaximum(len(self.filtered_cards) - 1)
            self.preview_idx = 0
            self.slider_nav.blockSignals(True)
            self.slider_nav.setValue(0)
            self.slider_nav.blockSignals(False)
            self.update_preview()
        else:
            self.lbl_preview.setText("V této kategorii nejsou žádné karty.")
            self.lbl_nav_total.setText("/ 0")
            self.inp_nav.setText("0")

    def get_dimensions_px(self):
        try:
            w_mm = float(self.inp_width.text().replace(',', '.'))
            h_mm = float(self.inp_height.text().replace(',', '.'))
            return int(w_mm * MM_TO_PX), int(h_mm * MM_TO_PX)
        except ValueError:
            return None, None

    def slider_changed(self, val):
        if not self.filtered_cards: return
        self.preview_idx = val
        self.update_preview()

    def input_changed(self):
        if not self.filtered_cards: return
        try:
            val = int(self.inp_nav.text()) - 1
            if 0 <= val < len(self.filtered_cards):
                self.preview_idx = val
                self.slider_nav.blockSignals(True)
                self.slider_nav.setValue(val)
                self.slider_nav.blockSignals(False)
                self.update_preview()
            else:
                self.inp_nav.setText(str(self.preview_idx + 1))
        except ValueError:
            self.inp_nav.setText(str(self.preview_idx + 1))

    def preview_prev(self):
        if self.filtered_cards and self.preview_idx > 0:
            self.preview_idx -= 1
            self.slider_nav.blockSignals(True)
            self.slider_nav.setValue(self.preview_idx)
            self.slider_nav.blockSignals(False)
            self.update_preview()

    def preview_next(self):
        if self.filtered_cards and self.preview_idx < len(self.filtered_cards) - 1:
            self.preview_idx += 1
            self.slider_nav.blockSignals(True)
            self.slider_nav.setValue(self.preview_idx)
            self.slider_nav.blockSignals(False)
            self.update_preview()

    def update_preview(self):
        if not self.filtered_cards: return
        w_px, h_px = self.get_dimensions_px()
        if w_px is None: return

        self.inp_nav.setText(str(self.preview_idx + 1))
        self.lbl_nav_total.setText(f"/ {len(self.filtered_cards)}")

        card = self.filtered_cards[self.preview_idx]
        pil_img = self.renderer.render(card, w_px, h_px)

        bytes_io = BytesIO()
        pil_img.save(bytes_io, format="PNG")
        pixmap = QPixmap()
        pixmap.loadFromData(bytes_io.getvalue())

        pw = self.lbl_preview.width()
        ph = self.lbl_preview.height()
        if pw > 20 and ph > 20:
            scaled_pixmap = pixmap.scaled(pw - 20, ph - 20, Qt.AspectRatioMode.KeepAspectRatio,
                                          Qt.TransformationMode.SmoothTransformation)
            self.lbl_preview.setPixmap(scaled_pixmap)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_preview()

    # ==========================================
    # 🚀 EXPORT VŠECH KARET
    # ==========================================
    def start_generation(self):
        if not self.parsed_cards:
            QMessageBox.warning(self, "Upozornění", "Nejprve načtěte data z Excelu.")
            return

        w_px, h_px = self.get_dimensions_px()
        if w_px is None:
            QMessageBox.critical(self, "Chyba", "Zadejte správně milimetry.")
            return

        self.btn_generate.setEnabled(False)
        self.btn_load_data.setEnabled(False)
        self.progress.setValue(0)

        # Worker zpracovává VŠECHNY karty, bez ohledu na filtr
        self.worker = CardGeneratorWorker(
            card_data_list=self.parsed_cards,
            dir_out=self.inp_dir_out.text(),
            excel_out=self.inp_excel_out.text(),
            width_px=w_px, height_px=h_px
        )
        self.worker.log_signal.connect(self.log_msg)
        self.worker.progress_signal.connect(self.progress.setValue)
        self.worker.finished_signal.connect(self.generation_finished)
        self.worker.start()

    def generation_finished(self, success, msg):
        self.progress.setValue(100)
        self.btn_generate.setEnabled(True)
        self.btn_load_data.setEnabled(True)
        if success:
            self.log_msg(f"\n✅ {msg}")
            QMessageBox.information(self, "Hotovo", msg)
        else:
            self.log_msg(f"\n❌ CHYBA: {msg}")
            QMessageBox.critical(self, "Chyba", msg)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = CardGeneratorApp()
    window.show()
    sys.exit(app.exec())